# report.py - Client Asset Report Generator with Financial Year Grouping
import os
import sys
import base64
import re
import traceback
from io import BytesIO
from datetime import date
import zipfile
import numpy as np

import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib.dates as mdates
import openpyxl
from matplotlib import font_manager as fm

from jinja2 import Template
from weasyprint import HTML

st.set_page_config(page_title="Client Asset Report Generator", layout="wide")
st.title("📊 Client Asset Report Generator (4-Page PDF)")

# ====== UPDATED TEMPLATE DIRECTORY ======
if getattr(sys, 'frozen', False):
    # Running as a PyInstaller bundle (.exe)
    TEMPLATE_DIR = sys._MEIPASS
else:
    # Running as normal .py script
    TEMPLATE_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------- Helpers ----------
def load_template(filename: str) -> Template:
    """Load an HTML file from TEMPLATE_DIR and return a Jinja2 Template."""
    path = os.path.join(TEMPLATE_DIR, filename)
    with open(path, "r", encoding="utf-8") as f:
        return Template(f.read())

def clean_number(x):
    """Coerce values like '₹ 1,23,456' or 'NA' to float safely."""
    try:
        x = str(x).replace("₹", "").replace(",", "").strip()
        return float(x) if x not in ["", "nan", "None"] else 0.0
    except Exception:
        return 0.0

def fig_to_base64_png(fig) -> str:
    """Convert a Matplotlib figure to base64 PNG string (no header)."""
    buf = BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=150, transparent=True)
    plt.close(fig)
    buf.seek(0)
    return base64.b64encode(buf.read()).decode("utf-8")

def format_indian_currency(amount):
    """Format number with Indian comma system (lakhs, crores) with ₹ symbol."""
    amount = float(amount)
    if amount == 0:
        return "₹ 0.00"
    
    amount_str = f"{amount:.2f}"
    
    if '.' in amount_str:
        integer_part, decimal_part = amount_str.split('.')
    else:
        integer_part, decimal_part = amount_str, "00"
    
    if len(integer_part) <= 3:
        formatted = integer_part
    else:
        last_three = integer_part[-3:]
        remaining = integer_part[:-3]
        
        groups = []
        while len(remaining) > 2:
            groups.append(remaining[-2:])
            remaining = remaining[:-2]
        if remaining:
            groups.append(remaining)
        
        groups.reverse()
        formatted = ','.join(groups) + ',' + last_three
    
    return f"₹ {formatted}.{decimal_part}"

def extract_client_name_from_filename(filename):
    """Extract client name from filename format: Portfolio_Computation_First_Last_Month.xlsx"""
    filename_no_ext = os.path.splitext(filename)[0]
    parts = filename_no_ext.split("_")
    
    if len(parts) >= 4:
        client_name_parts = parts[2:-1]
        client_name = " ".join(client_name_parts)
        return client_name.title()
    return "Client"

def get_appreciation_text(excel_file):
    """Get the appreciation text from Performance Report sheet, with amount + %."""
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Performance Report"]

        last_row = ws.max_row
        while not ws.cell(row=last_row, column=1).value:
            last_row -= 1

        absolute_appreciation = ws.cell(row=last_row, column=3).value  # Col C
        percentage_appreciation = ws.cell(row=last_row, column=4).value  # Col D

        abs_text = ""
        pct_text = ""

        if absolute_appreciation and absolute_appreciation != 0:
            if isinstance(absolute_appreciation, (int, float)):
                abs_text = format_indian_currency(absolute_appreciation)
            else:
                abs_text = str(absolute_appreciation)

        if percentage_appreciation and percentage_appreciation != 0:
            if isinstance(percentage_appreciation, (int, float)):
                pct_text = f"{percentage_appreciation:.2%}"
            else:
                pct_text = str(percentage_appreciation)

        if abs_text or pct_text:
           return (
                "<p style='font-style: italic; color: black; text-align: center;'>"
                "YOUR PORTFOLIO INVESTMENT VALUE HAS INCREASED BY "
                f"<span style='color:green;'>{abs_text}</span> OR "
                f"<span style='color:green;'>{pct_text}</span> "
                "ON ACCOUNT OF MARKET APPRECIATION."
                "</p>"
                )
        else:
            return "<i>YOUR PORTFOLIO INVESTMENT PERFORMANCE SUMMARY</i>"
            
    except Exception as e:
        st.error(f"Error reading appreciation data: {e}")
        return "<i>YOUR PORTFOLIO INVESTMENT PERFORMANCE SUMMARY</i>"

# (your generate_performance_chart, generate_comparison_chart, build_4page_html, build_client_pdf_bytes, and UI code remain unchanged)
# I didn’t touch anything else except TEMPLATE_DIR.


def generate_performance_chart(excel_file):
    """Generate performance line chart from Capital Contribution sheet with Y grid lines, grouping investments by financial year."""
    try:
        # Read both Amount Added and Amount Withdrawn columns
        cap_df = pd.read_excel(excel_file, sheet_name="Capital Contribution", usecols="A,B,D")
        cap_df.columns = ["Date", "Amount Added", "Amount Withdrawn"]
        
        # Handle NaN values - only fill NaN, keep actual 0 values as 0
        cap_df["Amount Added"] = cap_df["Amount Added"].apply(lambda x: 0 if pd.isna(x) else x)
        cap_df["Amount Withdrawn"] = cap_df["Amount Withdrawn"].apply(lambda x: 0 if pd.isna(x) else x)
        
        # Filter out rows where both Amount Added and Amount Withdrawn are 0
        cap_df = cap_df[(cap_df["Amount Added"] != 0) | (cap_df["Amount Withdrawn"] != 0)]
        
        cap_df["Date"] = pd.to_datetime(cap_df["Date"])
        
        # Group investments by financial year (April to March)
        def get_financial_year(dt):
            if dt.month >= 4:  # April to December
                return f"FY {dt.year}-{str(dt.year + 1)[-2:]}"
            else:  # January to March
                return f"FY {dt.year - 1}-{str(dt.year)[-2:]}"
        
        # Create financial year groups
        cap_df["Financial Year"] = cap_df["Date"].apply(get_financial_year)
        
        # Group by financial year and sum both added and withdrawn amounts
        grouped_df = cap_df.groupby("Financial Year").agg({
            "Date": "max",  # Use the last date in the financial year
            "Amount Added": "sum",  # Sum all investments in the financial year
            "Amount Withdrawn": "sum"  # Sum all withdrawals in the financial year
        }).reset_index()
        
        # Calculate net amount (added - withdrawn)
        grouped_df["Net Amount"] = grouped_df["Amount Added"] - grouped_df["Amount Withdrawn"]
        
        # Filter out financial years with no net activity
        grouped_df = grouped_df[grouped_df["Net Amount"] != 0]
        
        # Sort by date and calculate cumulative
        grouped_df = grouped_df.sort_values("Date")
        grouped_df["Cumulative"] = grouped_df["Net Amount"].cumsum()

        perf_df = pd.read_excel(excel_file, sheet_name="Performance Report", usecols="A:B")
        perf_df.columns = ["Date", "Portfolio Value"]
        perf_df["Date"] = pd.to_datetime(perf_df["Date"])
        latest_row = perf_df.iloc[-1]
        latest_date = latest_row["Date"]
        latest_value = latest_row["Portfolio Value"]

        # Combine grouped investments with latest portfolio value
        chart_df = grouped_df[["Date", "Cumulative"]].copy()
        chart_df = pd.concat([chart_df, pd.DataFrame({"Date": [latest_date], "Cumulative": [latest_value]})])
        chart_df = chart_df.sort_values("Date").reset_index(drop=True)

        # Set Open Sans font for matplotlib
        plt.rcParams['font.family'] = 'Open Sans'
        
        fig, ax = plt.subplots(figsize=(9, 6))
        
        # ✅ Add light transparent Y-axis grid lines
        ax.yaxis.grid(True, linestyle="--", alpha=0.3)

        # Filled area + line
        plt.fill_between(chart_df["Date"], chart_df["Cumulative"], color="#FFA366", alpha=0.9)
        plt.plot(chart_df["Date"], chart_df["Cumulative"], marker="o", color="#CC5500", linewidth=2)

        date_numeric = mdates.date2num(chart_df["Date"])
        x_min, x_max = date_numeric.min(), date_numeric.max()
        x_range = x_max - x_min
        
        y_min, y_max = ax.get_ylim()
        y_range = y_max - y_min
        
        # Create custom x-axis labels (FY labels for financial years, date for latest value)
        x_labels = []
        for i, row in chart_df.iterrows():
            if i < len(chart_df) - 1:  # Financial year entries
                fy_label = grouped_df.loc[grouped_df["Date"] == row["Date"], "Financial Year"].values[0]
                x_labels.append(fy_label)
            else:  # Latest portfolio value
                x_labels.append(row["Date"].strftime("%d-%b-%Y"))
        
        # Increase spacing between x-axis labels and center them properly
        plt.xticks(chart_df["Date"], x_labels, rotation=45, fontsize=9, ha='right')
        
        # Add extra space for the last label (portfolio value date)
        if len(chart_df) > 1:
            last_date_pos = mdates.date2num(chart_df["Date"].iloc[-1])
            second_last_date_pos = mdates.date2num(chart_df["Date"].iloc[-2])
            date_spacing = last_date_pos - second_last_date_pos
            
            # Extend x-axis limit to create more space for the last label
            ax.set_xlim([x_min - date_spacing * 0.2, x_max + date_spacing * 0.8])  # Adjusted for better centering
        
        for i, (x, y) in enumerate(zip(chart_df["Date"], chart_df["Cumulative"])):
            x_num = mdates.date2num(x)
            
            if i == len(chart_df) - 1:
                # Latest portfolio value
                plt.scatter(x, y, s=120, color="darkgreen", edgecolors="black", zorder=5)
                x_pos = mdates.num2date(x_num + (x_range * 0.02))
                plt.text(
                    x_pos, y + (y_range * 0.03),
                    f"₹{y:,.0f}",
                    ha="left",
                    fontsize=10, 
                    fontweight="bold",
                    color="#1E40AF",
                    bbox=dict(facecolor="white", edgecolor="#1E40AF", boxstyle="round,pad=0.3")
                )
            else:
                # Financial year NET amount (added - withdrawn)
                net_amount = grouped_df.loc[grouped_df["Date"] == x, "Net Amount"].values[0]
                
                # Determine arrow direction and color based on net amount
                if net_amount > 0:
                    # Net investment (deep blue arrow pointing UP from x-axis to line)
                    arrow_color = "#00008B"  # Deep blue
                    # Arrow from x-axis level to the data point
                    arrow_start_y = 0
                    arrow_end_y = y
                    amount_text = f"₹{net_amount:,.0f}"
                    text_y_pos = y_range * 0.12
                elif net_amount < 0:
                    # Net withdrawal (red arrow pointing DOWN from data point to x-axis)
                    arrow_color = "red"
                    # Arrow from data point to x-axis level
                    arrow_start_y = y
                    arrow_end_y = 0
                    amount_text = f"(₹{abs(net_amount):,.0f})"  # Negative amount in brackets
                    text_y_pos = y_range * 0.02
                else:
                    # No net change - skip this point
                    continue
                
                # Draw arrow line
                plt.plot([x, x], [arrow_start_y, arrow_end_y], color=arrow_color, alpha=0.9, linestyle="--", linewidth=2)
                
                # Add arrowhead
                arrow_direction = 1 if net_amount < 0 else -1
                plt.annotate("", 
                    xy=(x, arrow_end_y), 
                    xytext=(x, arrow_start_y),
                    arrowprops=dict(arrowstyle="->", color=arrow_color, lw=1.5)
                )
                
                # Add amount text
                x_text_pos = mdates.num2date(x_num + (x_range * 0.01))
                plt.text(
                    x_text_pos, text_y_pos,
                    amount_text,
                    ha="left",
                    fontsize=8, 
                    fontweight="bold", 
                    color=arrow_color
                )

        plt.gca().yaxis.set_major_formatter(mticker.StrMethodFormatter('{x:,.0f}'))
        plt.tick_params(axis="y", labelsize=9)
        plt.subplots_adjust(left=0.15, right=0.95, bottom=0.15)  # Adjusted to center the chart
        plt.gca().spines["top"].set_visible(False)
        plt.gca().spines["right"].set_visible(False)

        plt.tight_layout(pad=3.0)
        return fig_to_base64_png(fig)
        
    except Exception as e:
        st.error(f"Error generating performance chart: {e}")
        st.error(f"Detailed error: {traceback.format_exc()}")
        return ""

def generate_comparison_chart(excel_file):
    """Generate comparison bar chart from Performance Report sheet with two subplots: Absolute Return and Annualised Return."""
    try:
        # Load with openpyxl to get exact cell values
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb["Performance Report"]

        # Find last row with data in col A
        last_row = ws.max_row
        while not ws.cell(row=last_row, column=1).value:
            last_row -= 1

        # Get values for Absolute Return comparison
        pinkbull_value = ws.cell(row=last_row, column=2).value or 0  # Col B
        market_value = ws.cell(row=last_row, column=5).value or 0    # Col E

        # Get percentages for Absolute Return comparison
        pinkbull_pct_cell = ws.cell(row=last_row, column=4)  # Col D
        market_pct_cell = ws.cell(row=last_row, column=7)    # Col G

        pinkbull_pct_str = f"{pinkbull_pct_cell.value:.2%}" if isinstance(pinkbull_pct_cell.value, (int, float)) else str(pinkbull_pct_cell.value or "0%")
        market_pct_str = f"{market_pct_cell.value:.2%}" if isinstance(market_pct_cell.value, (int, float)) else str(market_pct_cell.value or "0%")

        # Get XIRR values for Annualised Return
        pinkbull_xirr = ws.cell(row=last_row, column=11).value or 0  # Col K
        market_xirr = ws.cell(row=last_row, column=12).value or 0    # Col L

        pinkbull_xirr_str = f"{pinkbull_xirr:.2%}" if isinstance(pinkbull_xirr, (int, float)) else str(pinkbull_xirr or "0%")
        market_xirr_str = f"{market_xirr:.2%}" if isinstance(market_xirr, (int, float)) else str(market_xirr or "0%")

        # Set Open Sans font for matplotlib
        plt.rcParams['font.family'] = 'Open Sans'
        
        # Create figure with two subplots side by side with increased space between them
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(9, 4.0))
        
        # Increase space between subplots even more
        plt.subplots_adjust(wspace=0.6)  # Increased from 0.4 to 0.6
        
        # First subplot: Absolute Return comparison
        x_labels_returns = ["MARKET*", "PINKBULL"]
        values_returns = [market_value, pinkbull_value]
        percentages_returns = [market_pct_str, pinkbull_pct_str]

        # Reduce space between individual bars and make them slimmer
        x_positions_returns = [0.5, 1.0]
        bars_returns = ax1.bar(
            x_positions_returns,
            values_returns,
            color="#FFA366",
            edgecolor="#CC5500",
            linewidth=1.5,
            width=0.3,  # Even narrower bars (from 0.4 to 0.3)
            zorder=3
        )

        # Add percentage text inside bars for Absolute Return
        for bar, pct_str in zip(bars_returns, percentages_returns):
            height = bar.get_height()
            ax1.text(
                bar.get_x() + bar.get_width() / 2,
                height * 0.85,
                pct_str,
                ha="center",
                va="top",
                fontsize=9,
                color="white",
                fontweight="bold"
            )

        ax1.set_xticks(x_positions_returns)
        ax1.set_xticklabels(x_labels_returns, fontsize=9)
        ax1.yaxis.set_major_formatter(mticker.StrMethodFormatter('{x:,.0f}'))
        ax1.spines["top"].set_visible(False)
        ax1.spines["right"].set_visible(False)
        ax1.grid(True, which='major', axis='y', linestyle=':', linewidth=0.7, color='grey', zorder=0)
        ax1.set_title('ABSOLUTE', fontsize=11, fontweight='bold', pad=15)
        ax1.tick_params(axis='y', labelsize=8)

        # Second subplot: Annualised Return (XIRR) comparison
        x_labels_xirr = ["MARKET*", "PINKBULL"]
        values_xirr = [market_xirr * 100 if isinstance(market_xirr, (int, float)) else 0, 
                      pinkbull_xirr * 100 if isinstance(pinkbull_xirr, (int, float)) else 0]
        percentages_xirr = [market_xirr_str, pinkbull_xirr_str]

        # Reduce space between individual bars and make them slimmer
        x_positions_xirr = [0.5, 1.0]
        bars_xirr = ax2.bar(
            x_positions_xirr,
            values_xirr,
            color="#FFA366",
            edgecolor="#CC5500",
            linewidth=1.5,
            width=0.3,  # Even narrower bars (from 0.4 to 0.3)
            zorder=3
        )

        # Add percentage text inside bars for Annualised Return
        for bar, pct_str in zip(bars_xirr, percentages_xirr):
            height = bar.get_height()
            ax2.text(
                bar.get_x() + bar.get_width() / 2,
                height * 0.85,
                pct_str,
                ha="center",
                va="top",
                fontsize=9,
                color="white",
                fontweight="bold"
            )

        ax2.set_xticks(x_positions_xirr)
        ax2.set_xticklabels(x_labels_xirr, fontsize=9)
        ax2.yaxis.set_major_formatter(mticker.PercentFormatter())
        ax2.spines["top"].set_visible(False)
        ax2.spines["right"].set_visible(False)
        ax2.grid(True, which='major', axis='y', linestyle=':', linewidth=0.7, color='grey', zorder=0)
        ax2.set_title('ANNUALISED', fontsize=11, fontweight='bold', pad=15)
        ax2.tick_params(axis='y', labelsize=8)

        # Remove excess whitespace and adjust layout
        plt.tight_layout(pad=2.0)
        
        return fig_to_base64_png(fig)
        
    except Exception as e:
        st.error(f"Error generating comparison chart: {e}")
        st.error(f"Detailed error: {traceback.format_exc()}")
        return ""

def build_4page_html(client_name: str, report_dt: date, perf_chart_b64: str, comp_chart_b64: str, appreciation_text: str, client_name_font_size: int = 45) -> str:
    """Build a 4-page unified HTML document."""
    
    client_name_upper = client_name.upper()
    report_date_str = report_dt.strftime("%d %B %Y").upper()  # Convert to uppercase
    
    # Create 4-page unified HTML with appreciation text
    unified_html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>Client Asset Report</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Open+Sans:ital,wght@0,300;0,400;0,600;0,700;0,800;1,300;1,400;1,600;1,700;1,800&display=swap');
  
  @page {{
    size: A4;
    margin: 0;
  }}
  
  body {{
    margin: 0;
    padding: 0;
    font-family: 'Open Sans', 'Arial', 'Helvetica', sans-serif;
  }}
  
  /* Page 1: Cover Page */
  .cover-page {{
    position: relative;
    width: 794px;
    height: 1123px;
    background-image: url('cover_page_bg.jpg');
    background-size: cover;
    background-repeat: no-repeat;
    page-break-after: always;
  }}
  
  .client-name {{
    position: absolute;
    top: 71px;
    left: 20px;
    font-size: {client_name_font_size}px;
    font-style: italic;
    color: white;
    font-family: 'Open Sans', sans-serif;
    font-weight: 600;
  }}
  
  .report-date {{
    position: absolute;
    top: 1032px;
    left: 420px;
    font-size: 30px;
    color: black;
    font-family: 'Open Sans', sans-serif;
    font-weight: 500;
    text-transform: uppercase;
  }}
  
  /* Page 2: Performance Page */
  .performance-page {{
    position: relative;
    width: 794px;
    height: 1123px;
    background-image: url('performance_page.jpg');
    background-size: cover;
    background-repeat: no-repeat;
    page-break-after: always;
  }}
  
  .performance-graph-area {{
    position: absolute;
    top: 42%;
    left: 50%;
    transform: translate(-50%, -50%);
    width: 780px;
    height: 480px;
    display: flex;
    justify-content: center;
    align-items: center;
  }}
  
  .performance-graph-area img {{
    width: 98%;
    height: 98%;
    object-fit: contain;
  }}
  
  .appreciation-text {{
    position: absolute;
    top: 75%;
    left: 50%;
    transform: translateX(-50%);
    width: 90%;
    text-align: center;
    font-size: 16px;
    font-weight: bold;
    font-style: italic;
    color: #228B22;
    background-color: rgba(255, 255, 255, 0.9);
    padding: 12px 20px;
    border-radius: 8px;
    box-shadow: 0 2px 6px rgba(0,0,0,0.1);
    z-index: 10;
    font-family: 'Open Sans', sans-serif;
  }}
  
  /* Page 3: Comparison Page */
  .comparison-page {{
    position: relative;
    width: 794px;
    height: 1123px;
    background-image: url('comparison.jpg');
    background-size: cover;
    background-repeat: no-repeat;
    display: flex;
    justify-content: center;
    align-items: center;
    page-break-after: always;
  }}
  
  .comparison-graph-area {{
    width: 900px;
    height: 350px;
    margin-top: -40px;
  }}
  
  .comparison-graph-area img {{
    width: 100%;
    height: 100%;
    object-fit: contain;
  }}
  
  /* Page 4: End Page */
  .end-page {{
    position: relative;
    width: 794px;
    height: 1123px;
    background-image: url('end_page_bg.jpg');
    background-size: cover;
    background-repeat: no-repeat;
  }}
</style>
</head>
<body>
  <!-- Page 1: Cover Page -->
  <div class="cover-page">
    <div class="client-name">{client_name_upper}</div>
    <div class="report-date">{report_date_str}</div>
  </div>
  
  <!-- Page 2: Performance Page -->
  <div class="performance-page">
    <div class="performance-graph-area">
      <img src="data:image/png;base64,{perf_chart_b64}" alt="Performance Graph">
    </div>
    <div class="appreciation-text">
      {appreciation_text}
    </div>
  </div>
  
  <!-- Page 3: Comparison Page -->
  <div class="comparison-page">
    <div class="comparison-graph-area">
      <img src="data:image/png;base64,{comp_chart_b64}" alt="Comparison Graph">
    </div>
  </div>
  
  <!-- Page 4: End Page -->
  <div class="end-page"></div>
</body>
</html>
"""
    return unified_html

def build_client_pdf_bytes(excel_file, client_name: str, report_dt: date, client_name_font_size: int = 45) -> bytes:
    """Generate 4-page PDF with performance and comparison charts."""
    
    # Generate both charts
    perf_chart_b64 = generate_performance_chart(excel_file)
    comp_chart_b64 = generate_comparison_chart(excel_file)
    
    # Get appreciation text
    appreciation_text = get_appreciation_text(excel_file)
    
    if not perf_chart_b64 or not comp_chart_b64:
        raise Exception("Failed to generate charts")
    
    # Build unified HTML document
    full_html = build_4page_html(client_name, report_dt, perf_chart_b64, comp_chart_b64, appreciation_text, client_name_font_size)
    
    # Generate PDF
    pdf_bytes = HTML(string=full_html, base_url=TEMPLATE_DIR).write_pdf()
    return pdf_bytes

# ---------- UI: inputs ----------
uploaded_file = st.file_uploader("Upload Client Excel File", type=["xlsx"], help="File should be named: Portfolio_Computation_Client_Name_Month.xlsx")
report_date = st.date_input("📅 Select Report Date", value=date.today())

# Client name font size control
st.write("### 🎨 Cover Page Styling")
client_name_font_size = st.slider(
    "Client Name Font Size (Cover Page)", 
    min_value=20, 
    max_value=80, 
    value=45, 
    step=1,
    help="Adjust the font size of the client name on the cover page"
)

st.caption(f"Templates folder: {TEMPLATE_DIR}")

# Add a separator
st.write("---")

if uploaded_file:
    try:
        # Extract client name from filename
        client_name = extract_client_name_from_filename(uploaded_file.name)
        st.success(f"✅ Client identified: {client_name}")
        
        # Show appreciation text preview
        appreciation_text = get_appreciation_text(uploaded_file)
        st.info(f"📈 Appreciation Text: {appreciation_text}")
        
        # Generate and download PDF
        pdf_bytes = build_client_pdf_bytes(uploaded_file, client_name, report_date, client_name_font_size)
        
        st.download_button(
            "📥 Download Client Report (PDF)",
            data=pdf_bytes,
            file_name=f"{client_name}_Asset_Report.pdf",
            mime="application/pdf",
            help="Download the 4-page PDF report with performance and comparison charts"
        )
        
    except Exception as e:
        st.error(f"Error generating report: {e}")
        st.error(f"Detailed error: {traceback.format_exc()}")
        st.info("Please ensure your Excel file has both 'Capital Contribution' and 'Performance Report' sheets with the expected format.")

